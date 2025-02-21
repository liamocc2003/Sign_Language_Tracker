import matplotlib.pyplot as plt
from openpyxl import load_workbook as load_wb


class PlotPointsFromHandToGraph():
    # Access Excel sheet for letter
    def get_points_from_excel(letter):
        list_of_points = []
        excel_file = r"Training NN\Datasets\Personal Dataset\personal_dataset.xlsx"

        #Load worksheet
        workbook = load_wb(excel_file)
        worksheet = workbook[letter]

        #Get all points
        column_count = worksheet.max_column - 1
        row_count = worksheet.max_row
        row_start = 2
        num_letters_in_alphabet = 26
        point = []
        all_points_for_letter = []

        for row_num in range(row_start, row_count):
            for column_num in range(column_count):
                if column_num > (num_letters_in_alphabet - 1):
                    column_num = (column_num % (num_letters_in_alphabet - 1) - 1)
                    ascii_for_column_letter = column_num + 65
                    column_letter = chr(ascii_for_column_letter)
                    column_letter = "A" + column_letter
                else:
                    ascii_for_column_letter = column_num + 65
                    column_letter = chr(ascii_for_column_letter)

                cell = column_letter + str(row_num)
                cell_value = worksheet[cell].value

                point.append(cell_value)

                if column_num % 2 == 1:
                    list_of_points.append(point)
                    point = []

            all_points_for_letter.append(list_of_points)
            list_of_points = []
            point = []
        
        return all_points_for_letter


    def plot_points(list_of_points):
        x_axis_index = 0
        y_axis_index = 1
        for i in range(len(list_of_points[0])): 
            x_axis = list_of_points[0][i][x_axis_index]
            y_axis = list_of_points[0][i][y_axis_index]    
            plt.plot(x_axis, y_axis, 'o')
        
        plt.gca().invert_yaxis()
        plt.show()

all_points = PlotPointsFromHandToGraph.get_points_from_excel("A")
print(all_points[0])

PlotPointsFromHandToGraph.plot_points(all_points)
