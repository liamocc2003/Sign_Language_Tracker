import time
#Get start timer for model
start_time = int(time.time())

import numpy as np
import os
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"

import matplotlib.pyplot as plt
from openpyxl import load_workbook as load_wb

from keras.models import Sequential
from keras import Input
from keras.layers import Dense, Dropout, Activation, Flatten, Convolution2D, MaxPooling2D
from keras.utils import to_categorical

class KerasDeepLearning():
    # Get data from Excel file
    def get_points_from_excel():

        excel_file = r"Training NN\Datasets\Personal Dataset\personal_dataset.xlsx"

        #Load worksheet
        workbook = load_wb(excel_file)

        ascii_for_starting_letter_A = 65
        num_letters_in_alphabet = 26
        ascii_for_final_letter_Z = ascii_for_starting_letter_A + num_letters_in_alphabet

        all_letters_and_their_points = []

        for letter_in_ascii in range(ascii_for_starting_letter_A, ascii_for_final_letter_Z):
            letter = chr(letter_in_ascii)
            
            letter_as_num = letter_in_ascii % ascii_for_starting_letter_A
            
            worksheet = workbook[letter]

            column_count = worksheet.max_column - 1
            row_count = worksheet.max_row + 1
            row_start = 2
            point = []
            list_of_points = []
            all_points_for_letter = []

            #Get all points
            for row_num in range(row_start, row_count):
                for column_num in range(column_count):
                    if column_num > (num_letters_in_alphabet - 1):
                        column_num = (column_num % (num_letters_in_alphabet - 1) - 1)
                        ascii_for_column_letter = column_num + 65
                        column_letter = chr(ascii_for_column_letter)
                        column_letter = "A" + column_letter
                    else:
                        ascii_for_column_letter = column_num + 65
                        column_letter = chr(ascii_for_column_letter)

                    cell = column_letter + str(row_num)
                    cell_value = worksheet[cell].value

                    point.append(cell_value)

                    if column_num % 2 == 1:
                        list_of_points.append(point)
                        point = []

                list_of_points.append(letter_as_num)
                all_points_for_letter.append(list_of_points)
                list_of_points = []
                point = []

            all_letters_and_their_points.append(all_points_for_letter)
        
        return all_letters_and_their_points

    # Split data into train and datasets
    def split_data_into_train_and_test(list_of_points):
        x_train = []
        x_test = []
        y_train = []
        y_test = []

        for letter in range(len(all_points_in_dataset)):
            for data_row in range(len(all_points_in_dataset[letter])):
                if data_row < (len(all_points_in_dataset[letter]) - 2):
                    x_train.append(all_points_in_dataset[letter][data_row][0:21])
                    y_train.append(all_points_in_dataset[letter][data_row][21])
                else:
                    x_test.append(all_points_in_dataset[letter][data_row][0:21])
                    y_test.append(all_points_in_dataset[letter][data_row][21])
        
        x_train = np.array(x_train)
        y_train = np.array(y_train)
        x_test = np.array(x_test)
        y_test = np.array(y_test)

        # print("Train Points:", x_train)
        # print("Train Letter:", y_train, "\nLength:", len(y_train))
        # print()
        # print("Test Points:", x_test)
        # print("Test Letter:", y_test, "\nLength:", len(y_test))

        return x_train, y_train, x_test, y_test


    # Preprocess data
    def preprocess_data(x_train, x_test):
        x_train = x_train.reshape(x_train.shape[0], 21, 2, 1)
        x_test = x_test.reshape(x_test.shape[0], 21, 2, 1)
        x_train = x_train.astype('float32')
        x_test = x_test.astype('float32')
        x_train /= 1000
        x_test /= 1000

        return x_train, x_test


    # Preprocess labels
    def preprocess_labels(y_train, y_test):
        num_letters_in_alphabet = 26
        
        y_train = to_categorical(y_train, num_letters_in_alphabet)
        y_test = to_categorical(y_test, num_letters_in_alphabet)

        return y_train, y_test


    # Model architecture
    def model_architecture():
        model = Sequential()
        model.add(Input((21, 2, 1)))

        model.add(Convolution2D(32, (2, 2), activation = 'relu'))

        model.add(Flatten())
        model.add(Dense(128, activation = 'relu'))
        model.add(Dense(64, activation = 'relu'))
        model.add(Dropout(0.2))
        model.add(Dense(26, activation='softmax'))
        
        return model


    # Compile model
    def compile_model(model):
        model.compile(loss = 'categorical_crossentropy', optimizer = 'adam', metrics = ['accuracy'])

        return model


    # Fit model on training data
    def fit_model(model, x_train, y_train, x_test, y_test, start_time):
        num_of_epochs = 200

        model.fit(x_train, y_train, batch_size = 8, epochs = num_of_epochs, verbose = 1)

        KerasDeepLearning.evaluate_model(model, x_test, y_test, start_time)

        return model


    # Evaluate model on testing data
    def evaluate_model(model, x_test, y_test, start_time):
        score = model.evaluate(x_test, y_test, verbose = 0)

        # Get total time model takes
        end_time = int(time.time())
        total_time_model_takes = end_time - start_time
        print()
        print("Model takes", total_time_model_takes , "seconds")

        print()
        print("Test values obtained using model:")
        print("Accuracy:", '{0:.2f}%'.format(score[1] * 100), "\nLoss:", '{0:.2f}%'.format(score[0] * 100))

        if (score[1] * 100) > 90:
            KerasDeepLearning.save_model(model)
            print()
            print("Model saved.")

        return score
    

    # Save model for future use
    def save_model(model):
        model.save("Training NN\keras_deep_learning_model.keras")


all_points_in_dataset = KerasDeepLearning.get_points_from_excel()
x_train, y_train, x_test, y_test = KerasDeepLearning.split_data_into_train_and_test(all_points_in_dataset)

x_train, x_test = KerasDeepLearning.preprocess_data(x_train, x_test)
y_train, y_test = KerasDeepLearning.preprocess_labels(y_train, y_test)

model = KerasDeepLearning.model_architecture()
model = KerasDeepLearning.compile_model(model)
model = KerasDeepLearning.fit_model(model, x_train, y_train, x_test, y_test, start_time)
